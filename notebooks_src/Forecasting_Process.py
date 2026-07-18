# Converted from: Forecasting Process.ipynb

# ===== Cell 0 =====
import pandas as pd
import glob
import os
from datetime import datetime
import numpy as np
import re
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report, accuracy_score

# ===== Cell 1 =====
# Cell 2: Load Historical Data
# Load the combined data with all features
all_data = pd.read_excel('Football_Data_With_Features.xlsx')

print(f"✓ Loaded historical data: {len(all_data)} matches")
print(f"\nAvailable Columns:")
for i, col in enumerate(all_data.columns, 1):
    print(f"{i:2d}. {col}")

print(f"\nSeasons available: {sorted(all_data['Season'].unique())}")
print(f"Rounds per season: {all_data.groupby('Season')['Round Number'].max().to_dict()}")

# ===== Cell 2 =====
# Cell 3: Separate Current Season from Historical (85/15 weighting)
# Identify current season (highest season number)
current_season = all_data['Season'].max()

current_season_df = all_data[all_data['Season'] == current_season].copy()
historical_df = all_data[all_data['Season'] < current_season].copy()

print(f"Current Season: Season {current_season}")
print(f"  - Matches: {len(current_season_df)}")
print(f"  - Rounds: {current_season_df['Round Number'].min()} to {current_season_df['Round Number'].max()}")

print(f"\nHistorical Seasons: {sorted(historical_df['Season'].unique())}")
print(f"  - Total Matches: {len(historical_df)}")

print(f"\nWeighting Strategy: 85% Current Season, 15% Historical")

# ===== Cell 3 =====
# Cell 3.5: Add Advanced Features to Training Data (FIXED)
print("="*80)
print("ADDING ADVANCED FEATURES TO TRAINING DATA")
print("="*80)

# ============================================================================
# 1. LAST 3 FORM % (Most recent form)
# ============================================================================
print("\n1. Calculating Last 3 Form % for all matches...")

all_data['Team A Last 3 Form %'] = 50.0  # Initialize with default
all_data['Team B Last 3 Form %'] = 50.0

def calculate_last3_form(team_data, current_idx, goals_for_col, goals_against_col):
    """Calculate form from last 3 games only"""
    # Get matches before current match
    previous_matches = team_data[team_data.index < current_idx].tail(3)
    
    if len(previous_matches) == 0:
        return 50.0  # Default if no history
    
    wins = sum(previous_matches[goals_for_col] > previous_matches[goals_against_col])
    draws = sum(previous_matches[goals_for_col] == previous_matches[goals_against_col])
    points = wins * 3 + draws
    
    return (points / (len(previous_matches) * 3)) * 100

# Calculate for each match
for idx, row in all_data.iterrows():
    team_a = row['Team A']
    team_b = row['Team B']
    
    # Team A last 3 form
    team_a_matches = all_data[all_data['Team A'] == team_a]
    if len(team_a_matches) > 0:
        all_data.at[idx, 'Team A Last 3 Form %'] = calculate_last3_form(
            team_a_matches, idx, 'Team A FT Goals', 'Team B FT Goals'
        )
    
    # Team B last 3 form
    team_b_matches = all_data[all_data['Team B'] == team_b]
    if len(team_b_matches) > 0:
        all_data.at[idx, 'Team B Last 3 Form %'] = calculate_last3_form(
            team_b_matches, idx, 'Team B FT Goals', 'Team A FT Goals'
        )

print(f"   ✓ Last 3 Form calculated for {len(all_data)} matches")

# ============================================================================
# 2. HOME/AWAY WIN % (Venue-specific performance)
# ============================================================================
print("\n2. Calculating Home/Away Win % for all teams...")

all_data['Team A Home Win %'] = 50.0  # Initialize
all_data['Team B Away Win %'] = 30.0

# Calculate for each unique team
for team in all_data['Team A'].unique():
    # Home performance (Team A only)
    team_home_matches = all_data[all_data['Team A'] == team].copy()
    
    if len(team_home_matches) > 0:
        home_wins = sum(team_home_matches['Team A FT Goals'] > team_home_matches['Team B FT Goals'])
        home_win_pct = (home_wins / len(team_home_matches)) * 100
        
        # Update all matches where this team is Team A
        all_data.loc[all_data['Team A'] == team, 'Team A Home Win %'] = home_win_pct
    
    # Away performance (Team B only)
    team_away_matches = all_data[all_data['Team B'] == team].copy()
    
    if len(team_away_matches) > 0:
        away_wins = sum(team_away_matches['Team B FT Goals'] > team_away_matches['Team A FT Goals'])
        away_win_pct = (away_wins / len(team_away_matches)) * 100
        
        # Update all matches where this team is Team B
        all_data.loc[all_data['Team B'] == team, 'Team B Away Win %'] = away_win_pct

print(f"   ✓ Home/Away splits calculated for {all_data['Team A'].nunique()} teams")

# ============================================================================
# 3. ODDS-BASED FEATURES (Skip for historical data - no odds available)
# ============================================================================
print("\n3. Handling Odds-based features...")

# Check if odds columns exist in historical data
if 'Home Win Odds' in all_data.columns:
    print("   → Odds data found, calculating odds features...")
    
    # Odds Favorite Margin
    all_data['Odds Favorite Margin'] = abs(
        (1 / all_data['Home Win Odds']) - (1 / all_data['Away Win Odds'])
    ) * 100
    
    # Clear Favorite flag
    all_data['Clear Favorite'] = (
        (all_data['Home Win Odds'] < 1.6) | 
        (all_data['Away Win Odds'] < 1.6)
    ).astype(int)
    
    # Draw Likelihood
    all_data['Draw Likelihood'] = (1 / all_data['Draw Odds']) * 100
    
    print(f"   ✓ Odds features calculated")
else:
    print("   ⚠️ No odds data in historical matches (expected)")
    print("   → Setting default values for training")
    
    # Set neutral default values for training
    all_data['Odds Favorite Margin'] = 10.0  # Moderate margin
    all_data['Clear Favorite'] = 0  # No clear favorite
    all_data['Draw Likelihood'] = 27.0  # ~3.7 odds

    print(f"   ✓ Odds features set to defaults (will use actual odds for predictions)")

# ============================================================================
# SUMMARY
# ============================================================================
print("\n" + "="*80)
print("ADVANCED FEATURES SUMMARY")
print("="*80)

print(f"\nTotal matches: {len(all_data)}")
print(f"Total features: {len(all_data.columns)}")

print("\nNew Features Added:")
print("   • Team A Last 3 Form % ✓")
print("   • Team B Last 3 Form % ✓")
print("   • Team A Home Win % ✓")
print("   • Team B Away Win % ✓")
print("   • Odds Favorite Margin ✓ (defaults for training)")
print("   • Clear Favorite ✓ (defaults for training)")
print("   • Draw Likelihood ✓ (defaults for training)")

print("\nSample Data:")
sample_cols = ['Team A', 'Team B', 'Team A Last 3 Form %', 'Team B Last 3 Form %', 
               'Team A Home Win %', 'Team B Away Win %']
print(all_data[sample_cols].tail(5))

# Verify no missing values
missing = all_data[['Team A Last 3 Form %', 'Team B Last 3 Form %', 
                     'Team A Home Win %', 'Team B Away Win %',
                     'Odds Favorite Margin', 'Clear Favorite', 'Draw Likelihood']].isnull().sum()

if missing.sum() > 0:
    print(f"\n⚠️ Warning: Found {missing.sum()} missing values")
    print(missing[missing > 0])
else:
    print("\n✓ No missing values in new features")

print("="*80)

# ===== Cell 4 =====
# Cell X: Define Round 12 Matches
round_12_matches = {
    'Team A': [
        'Z.Burnley', 'Z.Bournemouth', 'Z.Nottingham Forest', 'Z.Crystal Palace',
        'Z.Everton', 'Z.Chelsea', 'Z.West Ham Utd', 'Z.Newcastle Utd',
        'Z.Wolves', 'Z.Aston Villa'
    ],
    'Team B': [
        'Z.Arsenal', 'Z.Tottenham', 'Z.Sunderland', 'Z.Brentford',
        'Z.Man United', 'Z.Liverpool', 'Z.Fulham', 'Z.Leeds',
        'Z.Man City', 'Z.Brighton'
    ],
    'Home Win Odds': [
        8.65, 3.00, 1.50, 1.88,
        3.30, 3.55, 1.90, 1.40,
        7.80, 2.15
    ],
    'Draw Odds': [
        5.45, 3.50, 4.25, 3.55,
        3.60, 3.85, 3.55, 4.70,
        5.15, 3.45
    ],
    'Away Win Odds': [
        1.32, 2.27, 6.50, 4.15,
        2.08, 1.93, 4.00, 7.75,
        1.36, 3.25
    ],
    'Round Number': [12] * 10,
    'Round': ['Round 12'] * 10,
    'Match Number in Round': list(range(1, 11))
}

upcoming_matches = pd.DataFrame(round_12_matches)

print("Round 12 Matches to Forecast:")
print(
    upcoming_matches[
        ['Team A', 'Team B', 'Home Win Odds', 'Draw Odds', 'Away Win Odds']
    ]
)

# ===== Cell 5 =====
# Cell 5: Calculate ENHANCED Weighted Features for Upcoming Round
def get_team_form_weighted(team_name, current_data, historical_data, is_home=True):
    """Calculate team form with 85% current season, 15% historical weighting"""
    
    if is_home:
        current_matches = current_data[current_data['Team A'] == team_name].copy()
        hist_matches = historical_data[historical_data['Team A'] == team_name].copy()
        goals_for_col = 'Team A FT Goals'
        goals_against_col = 'Team B FT Goals'
    else:
        current_matches = current_data[current_data['Team B'] == team_name].copy()
        hist_matches = historical_data[historical_data['Team B'] == team_name].copy()
        goals_for_col = 'Team B FT Goals'
        goals_against_col = 'Team A FT Goals'
    
    # Get last 5 matches (for weighted average)
    current_last5 = current_matches.tail(5)
    hist_last5 = hist_matches.tail(5)
    
    # Get last 3 matches (for recent form)
    current_last3 = current_matches.tail(3)
    
    def calc_metrics(matches, gf_col, ga_col):
        if len(matches) == 0:
            return {'form_pct': 0, 'avg_scored': 0, 'avg_conceded': 0, 'games': 0}
        
        wins = sum(matches[gf_col] > matches[ga_col])
        draws = sum(matches[gf_col] == matches[ga_col])
        points = wins * 3 + draws
        
        return {
            'form_pct': (points / (len(matches) * 3)) * 100,
            'avg_scored': matches[gf_col].mean(),
            'avg_conceded': matches[ga_col].mean(),
            'games': len(matches)
        }
    
    current_metrics = calc_metrics(current_last5, goals_for_col, goals_against_col)
    hist_metrics = calc_metrics(hist_last5, goals_for_col, goals_against_col)
    last3_metrics = calc_metrics(current_last3, goals_for_col, goals_against_col)
    
    # Apply 85/15 weighting for main form
    if current_metrics['games'] > 0 and hist_metrics['games'] > 0:
        form_pct = (current_metrics['form_pct'] * 0.60) + (hist_metrics['form_pct'] * 0.40)
        avg_scored = (current_metrics['avg_scored'] * 0.60) + (hist_metrics['avg_scored'] * 0.40)
        avg_conceded = (current_metrics['avg_conceded'] * 0.60) + (hist_metrics['avg_conceded'] * 0.40)
    elif current_metrics['games'] > 0:
        form_pct = current_metrics['form_pct']
        avg_scored = current_metrics['avg_scored']
        avg_conceded = current_metrics['avg_conceded']
    elif hist_metrics['games'] > 0:
        form_pct = hist_metrics['form_pct']
        avg_scored = hist_metrics['avg_scored']
        avg_conceded = hist_metrics['avg_conceded']
    else:
        form_pct = 0
        avg_scored = 0
        avg_conceded = 0
    
    # Recent form (last 3 only)
    last3_form = last3_metrics['form_pct'] if last3_metrics['games'] > 0 else form_pct
    
    return {
        'form_pct': round(form_pct, 1),
        'avg_scored': round(avg_scored, 2),
        'avg_conceded': round(avg_conceded, 2),
        'current_games': current_metrics['games'],
        'current_form_pct': round(current_metrics['form_pct'], 1),
        'last3_form_pct': round(last3_form, 1)
    }

def get_home_away_splits(team_name, current_data, historical_data):
    """Calculate home and away win percentages"""
    
    # Home performance (as Team A)
    current_home = current_data[current_data['Team A'] == team_name]
    hist_home = historical_data[historical_data['Team A'] == team_name]
    all_home = pd.concat([current_home, hist_home])
    
    if len(all_home) > 0:
        home_wins = sum(all_home['Team A FT Goals'] > all_home['Team B FT Goals'])
        home_win_pct = (home_wins / len(all_home)) * 100
    else:
        home_win_pct = 50  # Default
    
    # Away performance (as Team B)
    current_away = current_data[current_data['Team B'] == team_name]
    hist_away = historical_data[historical_data['Team B'] == team_name]
    all_away = pd.concat([current_away, hist_away])
    
    if len(all_away) > 0:
        away_wins = sum(all_away['Team B FT Goals'] > all_away['Team A FT Goals'])
        away_win_pct = (away_wins / len(all_away)) * 100
    else:
        away_win_pct = 30  # Default
    
    return {
        'home_win_pct': round(home_win_pct, 1),
        'away_win_pct': round(away_win_pct, 1)
    }

def get_h2h_weighted(team_a, team_b, current_data, historical_data):
    """Calculate H2H with 85% current season, 15% historical weighting"""
    
    current_h2h = current_data[
        (((current_data['Team A'] == team_a) & (current_data['Team B'] == team_b)) |
         ((current_data['Team A'] == team_b) & (current_data['Team B'] == team_a)))
    ]
    
    hist_h2h = historical_data[
        (((historical_data['Team A'] == team_a) & (historical_data['Team B'] == team_b)) |
         ((historical_data['Team A'] == team_b) & (historical_data['Team B'] == team_a)))
    ]
    
    def calc_h2h(h2h_df, team_a):
        if len(h2h_df) == 0:
            return {'wins_a': 0, 'wins_b': 0, 'draws': 0, 'avg_goals': 0, 'total': 0}
        
        wins_a = 0
        wins_b = 0
        draws = 0
        
        for _, match in h2h_df.iterrows():
            if match['Team A'] == team_a:
                if match['Team A FT Goals'] > match['Team B FT Goals']:
                    wins_a += 1
                elif match['Team A FT Goals'] < match['Team B FT Goals']:
                    wins_b += 1
                else:
                    draws += 1
            else:
                if match['Team B FT Goals'] > match['Team A FT Goals']:
                    wins_a += 1
                elif match['Team B FT Goals'] < match['Team A FT Goals']:
                    wins_b += 1
                else:
                    draws += 1
        
        return {
            'wins_a': wins_a,
            'wins_b': wins_b,
            'draws': draws,
            'avg_goals': h2h_df['Total Goals'].mean(),
            'total': len(h2h_df)
        }
    
    current_h2h_metrics = calc_h2h(current_h2h, team_a)
    hist_h2h_metrics = calc_h2h(hist_h2h, team_a)
    
    total_matches = current_h2h_metrics['total'] + hist_h2h_metrics['total']
    
    if total_matches == 0:
        return {'matches': 0, 'team_a_wins': 0, 'team_a_win_pct': 0, 'avg_goals': 0}
    
    # Apply 85/15 weighting
    if current_h2h_metrics['total'] > 0 and hist_h2h_metrics['total'] > 0:
        wins_a = (current_h2h_metrics['wins_a'] * 0.60) + (hist_h2h_metrics['wins_a'] * 0.40)
        avg_goals = (current_h2h_metrics['avg_goals'] * 0.60) + (hist_h2h_metrics['avg_goals'] * 0.40)
    elif current_h2h_metrics['total'] > 0:
        wins_a = current_h2h_metrics['wins_a']
        avg_goals = current_h2h_metrics['avg_goals']
    else:
        wins_a = hist_h2h_metrics['wins_a']
        avg_goals = hist_h2h_metrics['avg_goals']
    
    return {
        'matches': total_matches,
        'team_a_wins': round(wins_a, 1),
        'team_a_win_pct': round((wins_a / total_matches) * 100, 1) if total_matches > 0 else 0,
        'avg_goals': round(avg_goals, 2)
    }

# ============================================================================
# CALCULATE FEATURES FOR UPCOMING MATCHES
# ============================================================================
print("Calculating ENHANCED weighted features for upcoming round matches...")

features_list = []
for idx, match in upcoming_matches.iterrows():
    team_a = match['Team A']
    team_b = match['Team B']
    
    # Get weighted team form
    team_a_form = get_team_form_weighted(team_a, current_season_df, historical_df, is_home=True)
    team_b_form = get_team_form_weighted(team_b, current_season_df, historical_df, is_home=False)
    
    # Get home/away splits
    team_a_splits = get_home_away_splits(team_a, current_season_df, historical_df)
    team_b_splits = get_home_away_splits(team_b, current_season_df, historical_df)
    
    # Get weighted H2H
    h2h = get_h2h_weighted(team_a, team_b, current_season_df, historical_df)
    
    # Calculate odds-based features
    home_odds = match['Home Win Odds']
    draw_odds = match['Draw Odds']
    away_odds = match['Away Win Odds']
    
    # Odds favorite margin
    odds_favorite_margin = abs((1/home_odds) - (1/away_odds)) * 100
    
    # Clear favorite flag
    clear_favorite = 1 if (home_odds < 1.6 or away_odds < 1.6) else 0
    
    # Draw likelihood
    draw_likelihood = (1 / draw_odds) * 100
    
    features = {
        'Team A': team_a,
        'Team B': team_b,
        'Round Number': match['Round Number'],
        'Match Number in Round': match['Match Number in Round'],
        'Home Win Odds': home_odds,
        'Draw Odds': draw_odds,
        'Away Win Odds': away_odds,
        
        # Team A - Standard
        'Team A Form %': team_a_form['form_pct'],
        'Team A Current Form %': team_a_form['current_form_pct'],
        'Team A Avg Scored': team_a_form['avg_scored'],
        'Team A Avg Conceded': team_a_form['avg_conceded'],
        'Team A Current Games': team_a_form['current_games'],
        
        # Team A - NEW: Recent + Home/Away
        'Team A Last 3 Form %': team_a_form['last3_form_pct'],
        'Team A Home Win %': team_a_splits['home_win_pct'],
        
        # Team B - Standard
        'Team B Form %': team_b_form['form_pct'],
        'Team B Current Form %': team_b_form['current_form_pct'],
        'Team B Avg Scored': team_b_form['avg_scored'],
        'Team B Avg Conceded': team_b_form['avg_conceded'],
        'Team B Current Games': team_b_form['current_games'],
        
        # Team B - NEW: Recent + Home/Away
        'Team B Last 3 Form %': team_b_form['last3_form_pct'],
        'Team B Away Win %': team_b_splits['away_win_pct'],
        
        # H2H
        'H2H Matches': h2h['matches'],
        'H2H Team A Wins': h2h['team_a_wins'],
        'H2H Team A Win %': h2h['team_a_win_pct'],
        'H2H Avg Goals': h2h['avg_goals'],
        
        # Derived
        'Form Difference': team_a_form['form_pct'] - team_b_form['form_pct'],
        
        # NEW: Odds-based features
        'Odds Favorite Margin': round(odds_favorite_margin, 1),
        'Clear Favorite': clear_favorite,
        'Draw Likelihood': round(draw_likelihood, 1)
    }
    
    features_list.append(features)

upcoming_features = pd.DataFrame(features_list)

print("✓ Enhanced features calculated\n")
print("Feature count:", len(upcoming_features.columns))
print("\nPreview (standard features):")
print(upcoming_features[['Team A', 'Team B', 'Team A Form %', 'Team B Form %', 'H2H Matches']].head())
print("\nPreview (new features):")
print(upcoming_features[['Team A', 'Team B', 'Team A Last 3 Form %', 'Team B Last 3 Form %', 
                         'Odds Favorite Margin', 'Clear Favorite']].head())

# ===== Cell 6 =====
# Cell 6: Enhanced ML Model Training - COMPLETE WITH CLASS BALANCING
import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split, GridSearchCV, cross_val_score
from sklearn.ensemble import RandomForestClassifier, VotingClassifier
from sklearn.metrics import classification_report, accuracy_score, confusion_matrix
from sklearn.preprocessing import LabelEncoder
from xgboost import XGBClassifier
import warnings
warnings.filterwarnings('ignore')

print("="*80)
print("ENHANCED ML MODEL TRAINING - WITH CLASS BALANCING")
print("="*80)

# ============================================================================
# 1. PREPARE TRAINING DATA
# ============================================================================
print("\n1. Preparing training data...")
training_data = all_data.dropna(subset=[
    'FT Result', 'Team A form_percentage', 'Team B form_percentage'
]).copy()

# Create sample weights (85/15 split - current season focus)
current_season_size = len(training_data[training_data['Season'] == current_season])
historical_size = len(training_data[training_data['Season'] < current_season])

sample_weights = np.where(
    training_data['Season'] == current_season,
    0.60 / current_season_size if current_season_size > 0 else 0,
    0.40 / historical_size if historical_size > 0 else 0
)
sample_weights = sample_weights * len(training_data) / sample_weights.sum()

print(f"   Total training samples: {len(training_data)}")
print(f"   Current season samples: {current_season_size} (60% weight)")
print(f"   Historical samples: {historical_size} (40% weight)")

feature_cols = [
    # Core form
    'Team A form_percentage', 'Team B form_percentage',
    'Team A avg_goals_scored', 'Team A avg_goals_conceded',
    'Team B avg_goals_scored', 'Team B avg_goals_conceded',
    
    # H2H
    'H2H Matches', 'H2H Team A Win %',
    
    # Context
    'Round Number', 'Match Number in Round',
    
    # NEW: Advanced features
    'Team A Last 3 Form %', 'Team B Last 3 Form %',
    'Team A Home Win %', 'Team B Away Win %',
    'Odds Favorite Margin', 'Clear Favorite', 'Draw Likelihood'
]

print(f"   Using {len(feature_cols)} features (7 new advanced features added)")

X = training_data[feature_cols]
y_raw = training_data['FT Result']

# Encode labels to numeric
label_encoder = LabelEncoder()
y = label_encoder.fit_transform(y_raw)

print(f"   Label encoding: {dict(zip(label_encoder.classes_, range(len(label_encoder.classes_))))}")

# Train/test split
X_train, X_test, y_train, y_test, w_train, w_test = train_test_split(
    X, y, sample_weights, test_size=0.2, random_state=42, stratify=y
)

print(f"   Training set: {len(X_train)} samples")
print(f"   Test set: {len(X_test)} samples")

# ============================================================================
# CHECK CLASS DISTRIBUTION (BEFORE TRAINING)
# ============================================================================
print("\n" + "="*80)
print("CLASS DISTRIBUTION ANALYSIS")
print("="*80)

y_train_labels = label_encoder.inverse_transform(y_train)
train_dist = pd.Series(y_train_labels).value_counts()
train_pct = pd.Series(y_train_labels).value_counts(normalize=True) * 100

print("\nTraining Data Distribution:")
for outcome in label_encoder.classes_:
    count = train_dist.get(outcome, 0)
    pct = train_pct.get(outcome, 0)
    print(f"   {outcome:12}: {count:4} ({pct:5.1f}%)")

home_pct = train_pct.get('Home Win', 0)
if home_pct > 45:
    print(f"\n⚠️ Home Win bias detected ({home_pct:.1f}%)")
    print("   → Using class_weight='balanced' to compensate")
elif home_pct < 30:
    print(f"\n⚠️ Away Win bias detected")
    print("   → Using class_weight='balanced' to compensate")
else:
    print(f"\n✓ Classes relatively balanced")

# ============================================================================
# MODEL 1: XGBoost with GridSearch + Class Balancing
# ============================================================================
print("\n" + "="*80)
print("2. Training XGBoost with GridSearch...")
print("="*80)

# Calculate class weights for XGBoost
from sklearn.utils.class_weight import compute_class_weight

class_weights = compute_class_weight(
    'balanced',
    classes=np.unique(y_train),
    y=y_train
)

# Create balanced sample weights
sample_weight_balanced = np.ones(len(y_train))
for i, class_label in enumerate(np.unique(y_train)):
    sample_weight_balanced[y_train == class_label] = class_weights[i]

# Combine with 85/15 temporal weighting
xgb_weights = w_train * sample_weight_balanced

xgb_param_grid = {
    'max_depth': [5, 7, 10],
    'learning_rate': [0.05, 0.1, 0.2],
    'n_estimators': [100, 200],
    'min_child_weight': [1, 3],
    'subsample': [0.8, 1.0],
    'colsample_bytree': [0.8, 1.0],
    'scale_pos_weight': [1, sum(y_train == 0) / sum(y_train == 2)]  # Balance classes
}

xgb_base = XGBClassifier(
    objective='multi:softmax',
    num_class=3,
    random_state=42,
    eval_metric='mlogloss',
    use_label_encoder=False
)

print("   Running GridSearch (this may take a few minutes)...")
xgb_grid = GridSearchCV(
    estimator=xgb_base,
    param_grid=xgb_param_grid,
    cv=3,
    scoring='accuracy',
    n_jobs=-1,
    verbose=0
)

xgb_grid.fit(X_train, y_train, sample_weight=xgb_weights)
xgb_model = xgb_grid.best_estimator_

print(f"   ✓ Best XGBoost parameters:")
for param, value in xgb_grid.best_params_.items():
    print(f"      {param}: {value}")

xgb_pred = xgb_model.predict(X_test)
xgb_accuracy = accuracy_score(y_test, xgb_pred)
print(f"\n   XGBoost Test Accuracy: {xgb_accuracy:.3f}")

# ============================================================================
# MODEL 2: Random Forest with Class Balancing
# ============================================================================
print("\n" + "="*80)
print("3. Training Random Forest with Class Balancing...")
print("="*80)

rf_param_grid = {
    'n_estimators': [100, 200, 300],
    'max_depth': [10, 15, 20, None],
    'min_samples_split': [2, 5, 10],
    'min_samples_leaf': [1, 2, 4],
    'class_weight': ['balanced', 'balanced_subsample', None]  # Test different balancing
}

rf_base = RandomForestClassifier(
    random_state=42,
    class_weight='balanced'  # Default to balanced
)

rf_grid = GridSearchCV(
    estimator=rf_base,
    param_grid=rf_param_grid,
    cv=3,
    scoring='accuracy',
    n_jobs=-1,
    verbose=0
)

rf_grid.fit(X_train, y_train, sample_weight=w_train)
rf_model = rf_grid.best_estimator_

print(f"   ✓ Best Random Forest parameters:")
for param, value in rf_grid.best_params_.items():
    print(f"      {param}: {value}")

rf_pred = rf_model.predict(X_test)
rf_accuracy = accuracy_score(y_test, rf_pred)
print(f"\n   Random Forest Test Accuracy: {rf_accuracy:.3f}")

# ============================================================================
# MODEL 3: Ensemble
# ============================================================================
print("\n" + "="*80)
print("4. Creating Ensemble Model...")
print("="*80)

ensemble_model = VotingClassifier(
    estimators=[
        ('xgb', xgb_model),
        ('rf', rf_model)
    ],
    voting='soft',
    weights=[2, 1]
)

# Combine weights for ensemble
ensemble_weights = w_train * sample_weight_balanced
ensemble_model.fit(X_train, y_train, sample_weight=ensemble_weights)

ensemble_pred = ensemble_model.predict(X_test)
ensemble_accuracy = accuracy_score(y_test, ensemble_pred)
print(f"   Ensemble Test Accuracy: {ensemble_accuracy:.3f}")

# ============================================================================
# MODEL COMPARISON
# ============================================================================
print("\n" + "="*80)
print("MODEL COMPARISON")
print("="*80)

results = pd.DataFrame({
    'Model': ['XGBoost', 'Random Forest', 'Ensemble'],
    'Accuracy': [xgb_accuracy, rf_accuracy, ensemble_accuracy]
}).sort_values('Accuracy', ascending=False)

print("\n" + results.to_string(index=False))

best_model_name = results.iloc[0]['Model']
best_accuracy = results.iloc[0]['Accuracy']

if best_model_name == 'XGBoost':
    final_model = xgb_model
elif best_model_name == 'Random Forest':
    final_model = rf_model
else:
    final_model = ensemble_model

print(f"\n🏆 Best Model: {best_model_name} ({best_accuracy:.3f} accuracy)")

# ============================================================================
# DETAILED EVALUATION
# ============================================================================
print("\n" + "="*80)
print(f"DETAILED EVALUATION - {best_model_name.upper()}")
print("="*80)

final_pred = final_model.predict(X_test)

# Decode labels for readable report
y_test_labels = label_encoder.inverse_transform(y_test)
final_pred_labels = label_encoder.inverse_transform(final_pred)

print("\nClassification Report:")
print(classification_report(y_test_labels, final_pred_labels))

print("\nConfusion Matrix:")
cm = confusion_matrix(y_test, final_pred)
cm_df = pd.DataFrame(
    cm,
    index=[f'True {label}' for label in label_encoder.classes_],
    columns=[f'Pred {label}' for label in label_encoder.classes_]
)
print(cm_df)

# Check prediction distribution on test set
print("\n" + "="*80)
print("PREDICTION DISTRIBUTION (Test Set)")
print("="*80)
pred_dist = pd.Series(final_pred_labels).value_counts()
pred_pct = pd.Series(final_pred_labels).value_counts(normalize=True) * 100

print("\nModel Predictions:")
for outcome in label_encoder.classes_:
    count = pred_dist.get(outcome, 0)
    pct = pred_pct.get(outcome, 0)
    print(f"   {outcome:12}: {count:4} ({pct:5.1f}%)")

# ============================================================================
# FEATURE IMPORTANCE
# ============================================================================
print("\n" + "="*80)
print("FEATURE IMPORTANCE")
print("="*80)

if best_model_name in ['XGBoost', 'Random Forest']:
    if hasattr(final_model, 'feature_importances_'):
        feature_importance = pd.DataFrame({
            'Feature': feature_cols,
            'Importance': final_model.feature_importances_
        }).sort_values('Importance', ascending=False)
        
        print("\nTop 5 Most Important Features:")
        print(feature_importance.head(5).to_string(index=False))

# ============================================================================
# CROSS-VALIDATION
# ============================================================================
print("\n" + "="*80)
print("CROSS-VALIDATION (Unweighted)")
print("="*80)

cv_scores = cross_val_score(
    final_model, X_train, y_train,
    cv=5,
    scoring='accuracy'
)

print(f"\n5-Fold CV Scores: {[f'{s:.3f}' for s in cv_scores]}")
print(f"Mean CV Accuracy: {cv_scores.mean():.3f} (+/- {cv_scores.std() * 2:.3f})")

# ============================================================================
# SAVE MODEL
# ============================================================================
print("\n" + "="*80)
print("SAVING MODEL")
print("="*80)

ml_model = final_model
model_name = best_model_name

print(f"\n✓ Best model ({model_name}) saved as 'ml_model'")
print(f"✓ Label encoder saved as 'label_encoder'")
print(f"✓ Feature columns saved as 'feature_cols'")

# ============================================================================
# FINAL SUMMARY
# ============================================================================
print("\n" + "="*80)
print("TRAINING SUMMARY")
print("="*80)
print(f"✓ Training completed successfully")
print(f"✓ Best Model: {model_name}")
print(f"✓ Test Accuracy: {best_accuracy:.3f}")
print(f"✓ CV Accuracy: {cv_scores.mean():.3f}")
print(f"✓ Class balancing: ENABLED")
print(f"✓ Sample weighting: 85% current / 15% historical")
print(f"✓ Ready for predictions!")
print("="*80)

# ===== Cell 7 =====
# Cell 7: ENHANCED Simple Prediction Function
def simple_prediction_v2(row):
    """Enhanced rule-based prediction with odds intelligence"""
    
    team_a_form = row['Team A Form %']
    team_b_form = row['Team B Form %']
    form_diff = row['Form Difference']
    
    home_odds = row['Home Win Odds']
    draw_odds = row['Draw Odds']
    away_odds = row['Away Win Odds']
    
    # Implied probabilities (remove bookmaker margin)
    total_prob = (1/home_odds + 1/draw_odds + 1/away_odds)
    home_true_prob = (1/home_odds / total_prob) * 100
    draw_true_prob = (1/draw_odds / total_prob) * 100
    away_true_prob = (1/away_odds / total_prob) * 100
    
    # ========================================================================
    # ENHANCED PREDICTION LOGIC
    # ========================================================================
    
    # 1. VERY STRONG FAVORITES (odds < 1.4)
    if home_odds < 1.4 and team_a_form > 55:
        pred = 'Home Win'
        conf = 'Very High'
        reason = 'Dominant Favorite'
        home_adj, draw_adj, away_adj = 85, 10, 5
    
    elif away_odds < 1.4 and team_b_form > 55:
        pred = 'Away Win'
        conf = 'Very High'
        reason = 'Dominant Favorite'
        home_adj, draw_adj, away_adj = 5, 10, 85
    
    # 2. STRONG FAVORITES (odds < 1.7)
    elif home_odds < 1.7 and form_diff > 15:
        pred = 'Home Win'
        conf = 'High'
        reason = 'Strong Favorite + Form'
        home_adj, draw_adj, away_adj = 70, 20, 10
    
    elif away_odds < 1.7 and form_diff < -15:
        pred = 'Away Win'
        conf = 'High'
        reason = 'Strong Favorite + Form'
        home_adj, draw_adj, away_adj = 10, 20, 70
    
    # 3. BIG FORM ADVANTAGE (>35% difference)
    elif form_diff > 35:
        pred = 'Home Win'
        conf = 'High'
        reason = 'Huge Form Advantage'
        home_adj, draw_adj, away_adj = 65, 25, 10
    
    elif form_diff < -35:
        pred = 'Away Win'
        conf = 'High'
        reason = 'Huge Form Advantage'
        home_adj, draw_adj, away_adj = 10, 25, 65
    
    # 4. MODERATE FORM ADVANTAGE (20-35%)
    elif form_diff > 20:
        pred = 'Home Win'
        conf = 'Medium'
        reason = 'Form Advantage'
        home_adj, draw_adj, away_adj = 55, 30, 15
    
    elif form_diff < -20:
        pred = 'Away Win'
        conf = 'Medium'
        reason = 'Form Advantage'
        home_adj, draw_adj, away_adj = 15, 30, 55
    
    # 5. EVENLY MATCHED - DRAW LIKELY
    elif abs(form_diff) < 12 and draw_odds < 3.8:
        pred = 'Draw'
        conf = 'Medium'
        reason = 'Evenly Matched'
        home_adj, draw_adj, away_adj = 30, 40, 30
    
    # 6. SLIGHT HOME ADVANTAGE
    elif form_diff > 0 and home_odds < away_odds:
        pred = 'Home Win'
        conf = 'Low'
        reason = 'Slight Edge'
        home_adj, draw_adj, away_adj = 45, 30, 25
    
    # 7. SLIGHT AWAY ADVANTAGE
    elif form_diff < 0 and away_odds < home_odds:
        pred = 'Away Win'
        conf = 'Low'
        reason = 'Slight Edge'
        home_adj, draw_adj, away_adj = 25, 30, 45
    
    # 8. DEFAULT TO ODDS FAVORITE
    else:
        if home_odds < away_odds:
            pred = 'Home Win'
        elif away_odds < home_odds:
            pred = 'Away Win'
        else:
            pred = 'Draw'
        conf = 'Low'
        reason = 'Odds Favorite'
        home_adj, draw_adj, away_adj = home_true_prob, draw_true_prob, away_true_prob
    
    return pd.Series({
        'Simple Prediction': pred,
        'Simple Confidence': conf,
        'Simple Reasoning': reason,
        'Simple Home Win %': round(home_adj, 1),
        'Simple Draw %': round(draw_adj, 1),
        'Simple Away Win %': round(away_adj, 1)
    })

print("✓ Enhanced simple prediction function defined")

# ===== Cell 8 =====
# Cell 8: Enhanced Betting Markets Prediction Function - COHERENT WITH MATCH RESULTS
def betting_markets_enhanced(row, simple_pred, ml_pred, ml_probs):
    """
    Predict betting markets intelligently based on match result predictions
    
    Args:
        row: Match data row
        simple_pred: Simple model prediction ('Home Win', 'Draw', 'Away Win')
        ml_pred: ML model prediction ('Home Win', 'Draw', 'Away Win')
        ml_probs: ML probability array [Away Win %, Draw %, Home Win %]
    """
    
    # ========================================================================
    # EXPECTED GOALS - Based on predicted result and team strength
    # ========================================================================
    team_a_attack = row['Team A Avg Scored']
    team_a_defense = row['Team A Avg Conceded']
    team_b_attack = row['Team B Avg Scored']
    team_b_defense = row['Team B Avg Conceded']
    
    # Base expected goals
    team_a_exp = (team_a_attack + team_b_defense) / 2
    team_b_exp = (team_b_attack + team_a_defense) / 2
    
    # Adjust based on ML prediction confidence
    home_win_prob = ml_probs[2]  # Home Win probability
    away_win_prob = ml_probs[0]  # Away Win probability
    draw_prob = ml_probs[1]      # Draw probability
    
    # If strong favorite, boost their expected goals
    if home_win_prob > 0.55:
        team_a_exp *= 1.2
        team_b_exp *= 0.85
    elif away_win_prob > 0.55:
        team_b_exp *= 1.2
        team_a_exp *= 0.85
    elif draw_prob > 0.35:
        # Draw likely = similar low scores
        team_a_exp *= 0.9
        team_b_exp *= 0.9
    
    total_exp_goals = team_a_exp + team_b_exp
    
    # ========================================================================
    # OVER/UNDER - Based on expected goals
    # ========================================================================
    over_1_5 = 'Yes' if total_exp_goals > 1.8 else 'No'
    over_2_5 = 'Yes' if total_exp_goals > 2.7 else 'No'
    over_3_5 = 'Yes' if total_exp_goals > 3.7 else 'No'
    
    # ========================================================================
    # BTTS - Smart logic based on both teams' attacking strength
    # ========================================================================
    # Strong team vs weak team = unlikely BTTS
    strength_diff = abs(row['Team A Form %'] - row['Team B Form %'])
    
    if strength_diff > 30:
        # Big mismatch = strong team scores, weak team doesn't
        btts = 'No'
    elif team_a_attack > 1.0 and team_b_attack > 1.0:
        # Both teams score regularly
        btts = 'Yes'
    elif team_a_attack < 0.7 or team_b_attack < 0.7:
        # One team struggles to score
        btts = 'No'
    else:
        # Medium scenario - check if it's a draw prediction
        btts = 'Yes' if draw_prob > 0.25 else 'No'
    
    # ========================================================================
    # LIKELY SCORE - Based on predicted result and expected goals
    # ========================================================================
    if ml_pred == 'Home Win':
        if total_exp_goals < 2.0:
            score = '1-0'
        elif total_exp_goals < 2.8:
            score = '2-0, 2-1'
        elif total_exp_goals < 3.5:
            score = '3-1, 2-1'
        else:
            score = '3-1, 4-1, 3-2'
    
    elif ml_pred == 'Away Win':
        if total_exp_goals < 2.0:
            score = '0-1'
        elif total_exp_goals < 2.8:
            score = '0-2, 1-2'
        elif total_exp_goals < 3.5:
            score = '1-3, 1-2'
        else:
            score = '1-3, 1-4, 2-3'
    
    else:  # Draw
        if total_exp_goals < 1.8:
            score = '0-0, 1-1'
        elif total_exp_goals < 2.8:
            score = '1-1, 2-2'
        else:
            score = '2-2, 3-3'
    
    # ========================================================================
    # HT/FT - Based on team strength and predicted result
    # ========================================================================
    form_diff = row['Team A Form %'] - row['Team B Form %']
    
    if ml_pred == 'Home Win':
        if home_win_prob > 0.65 and form_diff > 25:
            # Dominant home team
            ht_ft = 'Home Win/Home Win'
        elif form_diff > 15:
            # Good home team, careful start
            ht_ft = 'Draw/Home Win'
        else:
            # Tight game, home edges it
            ht_ft = 'Draw/Home Win'
    
    elif ml_pred == 'Away Win':
        if away_win_prob > 0.65 and form_diff < -25:
            # Dominant away team
            ht_ft = 'Away Win/Away Win'
        elif form_diff < -15:
            # Good away team
            ht_ft = 'Draw/Away Win'
        else:
            # Tight game, away edges it
            ht_ft = 'Draw/Away Win'
    
    else:  # Draw
        if draw_prob > 0.40:
            # Strong draw probability
            ht_ft = 'Draw/Draw'
        elif abs(form_diff) < 10:
            # Very evenly matched
            ht_ft = 'Draw/Draw'
        else:
            # Could go either way
            ht_ft = 'Draw/Draw'
    
    return pd.Series({
        'Expected Goals': round(total_exp_goals, 1),
        'Over 1.5': over_1_5,
        'Over 2.5': over_2_5,
        'Over 3.5': over_3_5,
        'BTTS': btts,
        'HT/FT': ht_ft,
        'Likely Score': score
    })

print("✓ Enhanced betting markets function defined")

# ===== Cell 9 =====
# Cell 9: Generate All Predictions (FIXED - All 17 Features)
print("Generating predictions...\n")

# Simple predictions
simple_results = upcoming_features.apply(simple_prediction_v2, axis=1)

# ============================================================================
# ML PREDICTIONS - WITH ALL 17 FEATURES
# ============================================================================
ml_features = upcoming_features[[
    # Core features (original 10)
    'Team A Form %', 'Team B Form %',
    'Team A Avg Scored', 'Team A Avg Conceded',
    'Team B Avg Scored', 'Team B Avg Conceded',
    'H2H Matches', 'H2H Team A Win %',
    'Round Number', 'Match Number in Round',
    
    # NEW: Advanced features (7 new ones)
    'Team A Last 3 Form %', 'Team B Last 3 Form %',
    'Team A Home Win %', 'Team B Away Win %',
    'Odds Favorite Margin', 'Clear Favorite', 'Draw Likelihood'
]].copy()

ml_features.columns = feature_cols

# Get numeric predictions and probabilities
ml_predictions_numeric = ml_model.predict(ml_features)
ml_probabilities = ml_model.predict_proba(ml_features)

# Decode numeric predictions to labels
ml_predictions = label_encoder.inverse_transform(ml_predictions_numeric)

# Get class labels
ml_classes = label_encoder.classes_  # ['Away Win', 'Draw', 'Home Win']

# Create ML probability dataframe with proper labels
ml_probs_df = pd.DataFrame(
    ml_probabilities * 100,
    columns=[f'ML {cls} %' for cls in ml_classes]
).round(1)

# ============================================================================
# BETTING MARKETS - ENHANCED VERSION WITH ML CONTEXT
# ============================================================================
print("Calculating betting markets with ML context...")

betting_results_list = []

for idx, row in upcoming_features.iterrows():
    # Get predictions for this match
    simple_pred = simple_results.loc[idx, 'Simple Prediction']
    ml_pred = ml_predictions[idx]
    ml_probs = ml_probabilities[idx]  # [Away Win prob, Draw prob, Home Win prob]
    
    # Calculate enhanced betting markets
    betting_result = betting_markets_enhanced(row, simple_pred, ml_pred, ml_probs)
    betting_results_list.append(betting_result)

betting_results = pd.DataFrame(betting_results_list, index=upcoming_features.index)

# ============================================================================
# COMBINE EVERYTHING
# ============================================================================
final_forecast = pd.concat([
    upcoming_features[['Team A', 'Team B', 'Home Win Odds', 'Draw Odds', 'Away Win Odds',
                       'Team A Form %', 'Team B Form %', 'Form Difference']],
    simple_results,
    pd.DataFrame({
        'ML Prediction': ml_predictions,
        'ML Confidence': (ml_probabilities.max(axis=1) * 100).round(1)
    }),
    ml_probs_df,
    betting_results
], axis=1)

print("✓ All predictions generated with coherent betting markets")

# ============================================================================
# SUMMARY STATISTICS
# ============================================================================
print("\n" + "="*80)
print("PREDICTION SUMMARY")
print("="*80)

print(f"\n🤖 ML Predictions:")
ml_summary = final_forecast['ML Prediction'].value_counts()
for pred, count in ml_summary.items():
    pct = (count / len(final_forecast)) * 100
    print(f"   {pred:12}: {count:2} ({pct:5.1f}%)")

print(f"\n🎯 Simple Predictions:")
simple_summary = final_forecast['Simple Prediction'].value_counts()
for pred, count in simple_summary.items():
    pct = (count / len(final_forecast)) * 100
    print(f"   {pred:12}: {count:2} ({pct:5.1f}%)")

print(f"\n⚽ Betting Markets:")
print(f"   Over 2.5 Yes: {(final_forecast['Over 2.5'] == 'Yes').sum()}/{len(final_forecast)}")
print(f"   BTTS Yes:     {(final_forecast['BTTS'] == 'Yes').sum()}/{len(final_forecast)}")

print(f"\n🤝 Model Agreement:")
agreement = (final_forecast['ML Prediction'] == final_forecast['Simple Prediction']).sum()
print(f"   {agreement}/{len(final_forecast)} matches ({(agreement/len(final_forecast)*100):.1f}%)")

print("="*80)

# ===== Cell 10 =====
# Cell 10: Display Match-by-Match
print("\n" + "="*120)
print("ROUND 28 FORECAST - MATCH BY MATCH")
print("="*120)

for idx, row in final_forecast.iterrows():
    print(f"\n{'─'*120}")
    print(f"⚽ Match {idx+1}: {row['Team A']} vs {row['Team B']}")
    print(f"{'─'*120}")
    
    print(f"\n🎯 PREDICTIONS:")
    print(f"   Odds: Home {row['Home Win Odds']} | Draw {row['Draw Odds']} | Away {row['Away Win Odds']}")
    print(f"")
    print(f"   Simple: {row['Simple Prediction']:12} ({row['Simple Confidence']})")
    print(f"           Home {row['Simple Home Win %']:.1f}% | Draw {row['Simple Draw %']:.1f}% | Away {row['Simple Away Win %']:.1f}%")
    print(f"           {row['Simple Reasoning']}")
    print(f"")
    print(f"   ML:     {row['ML Prediction']:12} ({row['ML Confidence']:.1f}%)")
    ml_prob_cols = [c for c in final_forecast.columns if 'ML ' in c and ' %' in c and c not in ['Simple Home Win %', 'Simple Draw %', 'Simple Away Win %']]
    ml_probs = " | ".join([f"{c.replace('ML ', '').replace(' %', '')} {row[c]:.1f}%" for c in ml_prob_cols])
    print(f"           {ml_probs}")
    
    print(f"\n⚽ BETTING MARKETS:")
    print(f"   Expected Goals: {row['Expected Goals']}")
    print(f"   Over 1.5: {row['Over 1.5']} | Over 2.5: {row['Over 2.5']} | Over 3.5: {row['Over 3.5']}")
    print(f"   BTTS: {row['BTTS']}")
    print(f"   HT/FT: {row['HT/FT']}")
    print(f"   Likely Score: {row['Likely Score']}")
    
    print(f"\n📈 FORM:")
    print(f"   {row['Team A']}: {row['Team A Form %']:.1f}%")
    print(f"   {row['Team B']}: {row['Team B Form %']:.1f}%")
    print(f"   Difference: {row['Form Difference']:+.1f}%")

print(f"\n{'='*120}\n")

# ===== Cell 11 =====
# Cell 11: Summary Statistics
print("="*120)
print("FORECAST SUMMARY")
print("="*120)

print(f"\n📊 Total Matches: {len(final_forecast)}")

print(f"\n🎯 SIMPLE PREDICTIONS:")
for outcome, count in final_forecast['Simple Prediction'].value_counts().items():
    pct = (count/len(final_forecast))*100
    print(f"   {outcome:12}: {count:2} ({pct:5.1f}%)")

print(f"\n🤖 ML PREDICTIONS:")
for outcome, count in final_forecast['ML Prediction'].value_counts().items():
    pct = (count/len(final_forecast))*100
    print(f"   {outcome:12}: {count:2} ({pct:5.1f}%)")

print(f"\n⚽ GOALS:")
for metric in ['Over 1.5', 'Over 2.5', 'Over 3.5']:
    print(f"   {metric}:")
    for val, cnt in final_forecast[metric].value_counts().items():
        pct = (cnt/len(final_forecast))*100
        print(f"      {val:3}: {cnt:2} ({pct:5.1f}%)")

print(f"\n🎯 BTTS:")
for val, cnt in final_forecast['BTTS'].value_counts().items():
    pct = (cnt/len(final_forecast))*100
    print(f"   {val:3}: {cnt:2} ({pct:5.1f}%)")

print(f"\n🎲 MODEL AGREEMENT:")
agreement = (final_forecast['Simple Prediction'] == final_forecast['ML Prediction']).sum()
print(f"   {agreement}/{len(final_forecast)} matches ({(agreement/len(final_forecast))*100:.1f}%)")

print(f"\n⚡ HIGH CONFIDENCE (ML ≥ 65%):")
high_conf = final_forecast[final_forecast['ML Confidence'] >= 65]
for idx, row in high_conf.iterrows():
    print(f"   {row['Team A']:20} vs {row['Team B']:20} → {row['ML Prediction']:12} ({row['ML Confidence']:.1f}%)")
print(f"\n{'='*120}\n")

# ===== Cell 12 =====
# Cell 12: Save Round 24 Predictions to Excel
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create a copy of final_forecast for saving
predictions_to_save = final_forecast.copy()

# Add timestamp
prediction_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
predictions_to_save['Prediction Time'] = prediction_time

# Add empty columns for actual results (to be filled in later)
predictions_to_save['Actual Result'] = ''
predictions_to_save['Actual Score'] = ''
predictions_to_save['Actual FT'] = ''
predictions_to_save['Actual HT'] = ''
predictions_to_save['Actual Total Goals'] = ''
predictions_to_save['Actual Over 2.5'] = ''
predictions_to_save['Actual BTTS'] = ''
predictions_to_save['Actual HT/FT'] = ''

# Add accuracy check columns (formulas will be added in Excel)
predictions_to_save['Simple Correct?'] = ''
predictions_to_save['ML Correct?'] = ''
predictions_to_save['Over 2.5 Correct?'] = ''
predictions_to_save['BTTS Correct?'] = ''

# Reorder columns for better organization
column_order = [
    # Match Info
    'Prediction Time',
    'Team A', 'Team B',
    'Home Win Odds', 'Draw Odds', 'Away Win Odds',
    
    # Predictions
    'Simple Prediction', 'Simple Confidence', 'Simple Reasoning',
    'ML Prediction', 'ML Confidence',
    
    # Betting Markets Predictions
    'Over 1.5', 'Over 2.5', 'Over 3.5',
    'BTTS', 'Expected Goals',
    'HT/FT', 'Likely Score',
    
    # Form Context
    'Team A Form %', 'Team B Form %', 'Form Difference',
    
    # Actual Results (Empty for now)
    'Actual Result', 'Actual Score', 
    'Actual FT', 'Actual HT',
    'Actual Total Goals', 'Actual Over 2.5', 
    'Actual BTTS', 'Actual HT/FT',
    
    # Accuracy Checks
    'Simple Correct?', 'ML Correct?', 
    'Over 2.5 Correct?', 'BTTS Correct?'
]

# Add ML probability columns
ml_prob_cols = [col for col in predictions_to_save.columns if 'ML ' in col and ' %' in col]
column_order = column_order[:12] + ml_prob_cols + column_order[12:]

# Select columns that exist
column_order = [col for col in column_order if col in predictions_to_save.columns]

predictions_to_save = predictions_to_save[column_order]

# Add match number
predictions_to_save.insert(0, 'Match #', range(1, len(predictions_to_save) + 1))

# Save basic Excel file first
filename = 'Round_12_Predictions_s13.xlsx'
predictions_to_save.to_excel(filename, index=False, sheet_name='Predictions')

print(f"✓ Saved basic predictions to: {filename}")

# Now enhance with formatting and formulas
wb = openpyxl.load_workbook(filename)
ws = wb['Predictions']

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
prediction_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
actual_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
accuracy_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Format headers
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Find column indices for formulas
col_names = [cell.value for cell in ws[1]]
simple_pred_col = col_names.index('Simple Prediction') + 1
ml_pred_col = col_names.index('ML Prediction') + 1
over25_pred_col = col_names.index('Over 2.5') + 1
btts_pred_col = col_names.index('BTTS') + 1

actual_result_col = col_names.index('Actual Result') + 1
actual_over25_col = col_names.index('Actual Over 2.5') + 1
actual_btts_col = col_names.index('Actual BTTS') + 1

simple_correct_col = col_names.index('Simple Correct?') + 1
ml_correct_col = col_names.index('ML Correct?') + 1
over25_correct_col = col_names.index('Over 2.5 Correct?') + 1
btts_correct_col = col_names.index('BTTS Correct?') + 1

# Add formulas for accuracy checks (starting from row 2)
for row in range(2, len(predictions_to_save) + 2):
    # Simple Correct?
    ws.cell(row, simple_correct_col).value = f'=IF({openpyxl.utils.get_column_letter(actual_result_col)}{row}="","",IF({openpyxl.utils.get_column_letter(simple_pred_col)}{row}={openpyxl.utils.get_column_letter(actual_result_col)}{row},"✓","✗"))'
    
    # ML Correct?
    ws.cell(row, ml_correct_col).value = f'=IF({openpyxl.utils.get_column_letter(actual_result_col)}{row}="","",IF({openpyxl.utils.get_column_letter(ml_pred_col)}{row}={openpyxl.utils.get_column_letter(actual_result_col)}{row},"✓","✗"))'
    
    # Over 2.5 Correct?
    ws.cell(row, over25_correct_col).value = f'=IF({openpyxl.utils.get_column_letter(actual_over25_col)}{row}="","",IF({openpyxl.utils.get_column_letter(over25_pred_col)}{row}={openpyxl.utils.get_column_letter(actual_over25_col)}{row},"✓","✗"))'
    
    # BTTS Correct?
    ws.cell(row, btts_correct_col).value = f'=IF({openpyxl.utils.get_column_letter(actual_btts_col)}{row}="","",IF({openpyxl.utils.get_column_letter(btts_pred_col)}{row}={openpyxl.utils.get_column_letter(actual_btts_col)}{row},"✓","✗"))'

# Color code columns
for row in range(2, len(predictions_to_save) + 2):
    # Prediction columns (gray)
    for col in range(simple_pred_col, actual_result_col):
        ws.cell(row, col).fill = prediction_fill
        ws.cell(row, col).border = thin_border
    
    # Actual result columns (yellow)
    for col in range(actual_result_col, simple_correct_col):
        ws.cell(row, col).fill = actual_fill
        ws.cell(row, col).border = thin_border
    
    # Accuracy columns (green)
    for col in range(simple_correct_col, len(col_names) + 1):
        ws.cell(row, col).fill = accuracy_fill
        ws.cell(row, col).border = thin_border
        ws.cell(row, col).alignment = Alignment(horizontal='center')

# Add summary section at the bottom
summary_row = len(predictions_to_save) + 3

ws.cell(summary_row, 1).value = "SUMMARY STATISTICS"
ws.cell(summary_row, 1).font = Font(bold=True, size=12)

summary_row += 1
ws.cell(summary_row, 1).value = "Total Matches:"
ws.cell(summary_row, 2).value = len(predictions_to_save)

summary_row += 1
ws.cell(summary_row, 1).value = "Simple Accuracy:"
ws.cell(summary_row, 2).value = f'=COUNTIF({openpyxl.utils.get_column_letter(simple_correct_col)}2:{openpyxl.utils.get_column_letter(simple_correct_col)}{len(predictions_to_save)+1},"✓")/COUNTA({openpyxl.utils.get_column_letter(actual_result_col)}2:{openpyxl.utils.get_column_letter(actual_result_col)}{len(predictions_to_save)+1})'
ws.cell(summary_row, 2).number_format = '0.0%'

summary_row += 1
ws.cell(summary_row, 1).value = "ML Accuracy:"
ws.cell(summary_row, 2).value = f'=COUNTIF({openpyxl.utils.get_column_letter(ml_correct_col)}2:{openpyxl.utils.get_column_letter(ml_correct_col)}{len(predictions_to_save)+1},"✓")/COUNTA({openpyxl.utils.get_column_letter(actual_result_col)}2:{openpyxl.utils.get_column_letter(actual_result_col)}{len(predictions_to_save)+1})'
ws.cell(summary_row, 2).number_format = '0.0%'

summary_row += 1
ws.cell(summary_row, 1).value = "Over 2.5 Accuracy:"
ws.cell(summary_row, 2).value = f'=COUNTIF({openpyxl.utils.get_column_letter(over25_correct_col)}2:{openpyxl.utils.get_column_letter(over25_correct_col)}{len(predictions_to_save)+1},"✓")/COUNTA({openpyxl.utils.get_column_letter(actual_over25_col)}2:{openpyxl.utils.get_column_letter(actual_over25_col)}{len(predictions_to_save)+1})'
ws.cell(summary_row, 2).number_format = '0.0%'

summary_row += 1
ws.cell(summary_row, 1).value = "BTTS Accuracy:"
ws.cell(summary_row, 2).value = f'=COUNTIF({openpyxl.utils.get_column_letter(btts_correct_col)}2:{openpyxl.utils.get_column_letter(btts_correct_col)}{len(predictions_to_save)+1},"✓")/COUNTA({openpyxl.utils.get_column_letter(actual_btts_col)}2:{openpyxl.utils.get_column_letter(actual_btts_col)}{len(predictions_to_save)+1})'
ws.cell(summary_row, 2).number_format = '0.0%'

# Adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column_letter].width = adjusted_width

# Freeze first row and first two columns
ws.freeze_panes = 'C2'

# Save the formatted workbook
wb.save(filename)

print(f"\n{'='*80}")
print(f"✓ PREDICTIONS SAVED WITH FORMATTING")
print(f"{'='*80}")
print(f"File: {filename}")
print(f"Matches: {len(predictions_to_save)}")
print(f"\nNext steps:")
print(f"1. Keep this file safe")
print(f"2. After Round 5 finishes, open the file")
print(f"3. Fill in the yellow 'Actual' columns")
print(f"4. Green checkmarks/crosses will auto-calculate")
print(f"5. Summary statistics will auto-update")
print(f"{'='*80}\n")

# ===== Cell 13 =====
# Cell: Update Predictions WITHOUT Destroying Formatting
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
import re

print("="*80)
print("UPDATING ROUND 29 PREDICTIONS (PRESERVING FORMATTING)")
print("="*80)

# Configuration
round_number = 12
predictions_file = 'Round_12_Predictions_s13.xlsx'
actuals_file = 'All_Rounds_Combined_s13.xlsx'

# Helper functions
def extract_goals(score_str):
    """Extract goals from score like '2-0' or '1-0HT'"""
    if pd.isna(score_str):
        return None, None
    score_str = str(score_str).replace('HT', '').strip()
    parts = score_str.split('-')
    if len(parts) == 2:
        try:
            return int(parts[0]), int(parts[1])
        except:
            return None, None
    return None, None

def get_ft_result(team_a_goals, team_b_goals):
    """Get FT Result from goals"""
    if pd.isna(team_a_goals) or pd.isna(team_b_goals):
        return None
    if team_a_goals > team_b_goals:
        return 'Home Win'
    elif team_a_goals < team_b_goals:
        return 'Away Win'
    else:
        return 'Draw'

# Step 1: Load actual results
print(f"\n1. Loading actual results from: {actuals_file}")
try:
    actuals_df = pd.read_excel(actuals_file)
    actuals_df = actuals_df[actuals_df['Round Number'] == round_number]
    print(f"   ✓ Found {len(actuals_df)} matches for Round {round_number}")
except Exception as e:
    print(f"   ✗ Error: {e}")
    exit()

# Calculate features
actuals_df[['Team A FT Goals', 'Team B FT Goals']] = actuals_df['FT Score'].apply(
    lambda x: pd.Series(extract_goals(x))
)
actuals_df[['Team A HT Goals', 'Team B HT Goals']] = actuals_df['HT Score'].apply(
    lambda x: pd.Series(extract_goals(x))
)
actuals_df['FT Result'] = actuals_df.apply(
    lambda row: get_ft_result(row['Team A FT Goals'], row['Team B FT Goals']), axis=1
)
actuals_df['Total Goals'] = actuals_df['Team A FT Goals'] + actuals_df['Team B FT Goals']
actuals_df['Over 2.5'] = actuals_df['Total Goals'] > 2.5
actuals_df['BTTS'] = (actuals_df['Team A FT Goals'] > 0) & (actuals_df['Team B FT Goals'] > 0)
actuals_df['HT Result'] = actuals_df.apply(
    lambda row: get_ft_result(row['Team A HT Goals'], row['Team B HT Goals']), axis=1
)
actuals_df['HT/FT Pattern'] = actuals_df['HT Result'] + '/' + actuals_df['FT Result']

print(f"   ✓ Features calculated")

# Step 2: Open existing Excel file (DON'T overwrite!)
print(f"\n2. Opening existing predictions file: {predictions_file}")
try:
    wb = openpyxl.load_workbook(predictions_file)
    ws = wb['Predictions']
    print(f"   ✓ Opened existing file")
except Exception as e:
    print(f"   ✗ Error: {e}")
    exit()

# Step 3: Find column indices
col_names = [cell.value for cell in ws[1]]

team_a_col = col_names.index('Team A') + 1
team_b_col = col_names.index('Team B') + 1
simple_pred_col = col_names.index('Simple Prediction') + 1
ml_pred_col = col_names.index('ML Prediction') + 1

# Actual columns (yellow - we'll fill these)
actual_result_col = col_names.index('Actual Result') + 1
actual_score_col = col_names.index('Actual Score') + 1
actual_ft_col = col_names.index('Actual FT') + 1
actual_ht_col = col_names.index('Actual HT') + 1
actual_total_goals_col = col_names.index('Actual Total Goals') + 1
actual_over25_col = col_names.index('Actual Over 2.5') + 1
actual_btts_col = col_names.index('Actual BTTS') + 1
actual_htft_col = col_names.index('Actual HT/FT') + 1

print(f"   ✓ Column indices found")

# Step 4: Update ONLY the Actual columns (row by row)
print(f"\n3. Updating actual results...")

matches_found = 0
matches_not_found = []

# Start from row 2 (after headers)
for row_idx in range(2, ws.max_row + 1):
    team_a = ws.cell(row_idx, team_a_col).value
    team_b = ws.cell(row_idx, team_b_col).value
    
    # Skip empty rows
    if pd.isna(team_a) or pd.isna(team_b) or team_a == '' or team_b == '':
        continue
    
    # Find matching actual result
    actual_match = actuals_df[
        (actuals_df['Team A'] == team_a) & 
        (actuals_df['Team B'] == team_b)
    ]
    
    if len(actual_match) > 0:
        actual = actual_match.iloc[0]
        
        # Fill in ONLY the yellow "Actual" columns
        ws.cell(row_idx, actual_result_col).value = actual['FT Result']
        ws.cell(row_idx, actual_score_col).value = actual['FT Score']
        ws.cell(row_idx, actual_ft_col).value = actual['FT Score']
        ws.cell(row_idx, actual_ht_col).value = actual['HT Score']
        ws.cell(row_idx, actual_total_goals_col).value = int(actual['Total Goals'])
        ws.cell(row_idx, actual_over25_col).value = 'Yes' if actual['Over 2.5'] else 'No'
        ws.cell(row_idx, actual_btts_col).value = 'Yes' if actual['BTTS'] else 'No'
        ws.cell(row_idx, actual_htft_col).value = actual['HT/FT Pattern']
        
        matches_found += 1
        print(f"   ✓ Row {row_idx-1}: {team_a:20} vs {team_b:20}: {actual['FT Result']:12} ({actual['FT Score']})")
    else:
        matches_not_found.append(f"{team_a} vs {team_b}")
        print(f"   ✗ Row {row_idx-1}: {team_a} vs {team_b}: NOT FOUND")

print(f"\n   Summary: {matches_found} matches updated")

if matches_not_found:
    print(f"   ⚠️ Not found: {matches_not_found}")

# Step 5: Color code the checkmarks (green/red)
print(f"\n4. Updating checkmark colors...")

simple_correct_col = col_names.index('Simple Correct?') + 1
ml_correct_col = col_names.index('ML Correct?') + 1
over25_correct_col = col_names.index('Over 2.5 Correct?') + 1
btts_correct_col = col_names.index('BTTS Correct?') + 1

correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
incorrect_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

for row_idx in range(2, ws.max_row + 1):
    # Skip if no actual result
    if not ws.cell(row_idx, actual_result_col).value:
        continue
    
    # Color Simple Correct
    cell = ws.cell(row_idx, simple_correct_col)
    if cell.value == '✓':
        cell.fill = correct_fill
        cell.font = Font(color="006100", bold=True, size=12)
    elif cell.value == '✗':
        cell.fill = incorrect_fill
        cell.font = Font(color="9C0006", bold=True, size=12)
    
    # Color ML Correct
    cell = ws.cell(row_idx, ml_correct_col)
    if cell.value == '✓':
        cell.fill = correct_fill
        cell.font = Font(color="006100", bold=True, size=12)
    elif cell.value == '✗':
        cell.fill = incorrect_fill
        cell.font = Font(color="9C0006", bold=True, size=12)
    
    # Color Over 2.5 Correct
    cell = ws.cell(row_idx, over25_correct_col)
    if cell.value == '✓':
        cell.fill = correct_fill
        cell.font = Font(color="006100", bold=True, size=12)
    elif cell.value == '✗':
        cell.fill = incorrect_fill
        cell.font = Font(color="9C0006", bold=True, size=12)
    
    # Color BTTS Correct
    cell = ws.cell(row_idx, btts_correct_col)
    if cell.value == '✓':
        cell.fill = correct_fill
        cell.font = Font(color="006100", bold=True, size=12)
    elif cell.value == '✗':
        cell.fill = incorrect_fill
        cell.font = Font(color="9C0006", bold=True, size=12)

print(f"   ✓ Checkmarks color-coded")

# Step 6: Save (preserving all formatting)
print(f"\n5. Saving updated file...")
wb.save(predictions_file)
print(f"   ✓ File saved: {predictions_file}")

# Step 7: Calculate and display accuracy
print(f"\n{'='*80}")
print(f"ROUND {round_number} RESULTS")
print(f"{'='*80}")

simple_correct = 0
ml_correct = 0
over25_correct = 0
btts_correct = 0
total_matches = 0

for row_idx in range(2, ws.max_row + 1):
    actual_result = ws.cell(row_idx, actual_result_col).value
    if not actual_result or actual_result == '':
        continue
    
    total_matches += 1
    
    simple_pred = ws.cell(row_idx, simple_pred_col).value
    ml_pred = ws.cell(row_idx, ml_pred_col).value
    
    if simple_pred == actual_result:
        simple_correct += 1
    if ml_pred == actual_result:
        ml_correct += 1
    
    # Check formulas for accuracy
    if ws.cell(row_idx, over25_correct_col).value == '✓':
        over25_correct += 1
    if ws.cell(row_idx, btts_correct_col).value == '✓':
        btts_correct += 1

simple_accuracy = (simple_correct / total_matches * 100) if total_matches > 0 else 0
ml_accuracy = (ml_correct / total_matches * 100) if total_matches > 0 else 0
over25_accuracy = (over25_correct / total_matches * 100) if total_matches > 0 else 0
btts_accuracy = (btts_correct / total_matches * 100) if total_matches > 0 else 0

print(f"\nOVERALL ACCURACY:")
print(f"{'-'*80}")
print(f"Simple Model:  {simple_correct:2}/{total_matches:2} ({simple_accuracy:5.1f}%)")
print(f"ML Model:      {ml_correct:2}/{total_matches:2} ({ml_accuracy:5.1f}%)")
print(f"Over 2.5:      {over25_correct:2}/{total_matches:2} ({over25_accuracy:5.1f}%)")
print(f"BTTS:          {btts_correct:2}/{total_matches:2} ({btts_accuracy:5.1f}%)")

if simple_accuracy > ml_accuracy:
    print(f"\n🏆 Simple Model won by {simple_accuracy - ml_accuracy:.1f}%")
elif ml_accuracy > simple_accuracy:
    print(f"\n🏆 ML Model won by {ml_accuracy - simple_accuracy:.1f}%")
else:
    print(f"\n🤝 Both models tied")

print(f"\n{'='*20}")
print(f"✓ Round {round_number} updated successfully!")
print(f"✓ Open {predictions_file} to see results")
print(f"✓ All formatting preserved!")
print(f"{'='*20}\n")
